import os
import math
import logging
import chardet
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import argparse
import pandas as pd
from datetime import datetime

 
LOGS_DIRNAME = 'logs'

SPECIMEN_FINAL      = 'Specimen Final'
SPECIMEN            = 'Specimen'
ORIGINATOR          = 'Originator'
SSTI                = 'SSTI'
BLOOD               = 'Blood'
RESULTS             = 'Results'
MRSA                = 'MRSA'
MSSA                = 'MSSA'
PVL                 = 'RT-pvl Result'
POSITIVE            = 'Positive'
NEGATIVE            = 'Negative'

MRSA_POSITIVE_SSTI      = 'MRSA PVL POSITIVE SSTI'
MRSA_POSITIVE_BLOOD     = 'MRSA PVL POSITIVE BLOOD'
MRSA_NEGATIVE_SSTI      = 'MRSA PVL NEGATIVE SSTI'
MRSA_NEGATIVE_BLOOD     = 'MRSA PVL NEGATIVE BLOOD'

ORANGE_FILL = PatternFill(start_color="FFB732", end_color="FFB732", fill_type="solid")
GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid") 
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
TITLES = [MRSA_POSITIVE_SSTI, MRSA_POSITIVE_BLOOD, MRSA_NEGATIVE_SSTI, MRSA_NEGATIVE_BLOOD, MSSA]

 
def setup_logs():
    logs_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), LOGS_DIRNAME)
    timestamp = datetime.now().strftime("%y%m%d_%H%M%S")
    logs_file = os.path.join(logs_dir, f'logs__{timestamp}.txt')

    os.makedirs(logs_dir, exist_ok=True)

    logging.basicConfig(filename=logs_file,level=logging.DEBUG,format="%(message)s")
    
def display_and_log(msg, is_error=False):
    print(f'{msg}\n')
    if is_error:
        logging.error(f'{msg}\n')
    else:
        logging.info(f'{msg}\n')

 
def load_data(input_file, sheetname):
    if input_file.endswith(".csv"):
        with open(input_file, 'rb') as f:
            result = chardet.detect(f.read(100000))  # Lire une partie du fichier
            detected_encoding = result['encoding']
        return pd.read_csv(input_file, encoding=detected_encoding)  # Ensure Hebrew support

    elif input_file.endswith(".xls") or input_file.endswith(".xlsx"):
        if input_file.endswith(".xls") or input_file.endswith(".xlsx"):

            return pd.read_excel(input_file, sheet_name=sheetname)

        else:

            raise ValueError("Unsupported file format. Please provide CSV or Excel file.")

def order_the_table(df):
    final_df = pd.DataFrame()

    def add_section(title, subset_df):
        """Helper function to add a title, column headers, the data, and an empty row."""

        nonlocal final_df

        if not subset_df.empty:
            title_row = pd.DataFrame([[title] + [''] * (len(df.columns) - 1)], columns=df.columns)
            header_row = pd.DataFrame([df.columns.tolist()], columns=df.columns)
            empty_row = pd.DataFrame([[''] * len(df.columns)], columns=df.columns)

            final_df = pd.concat([final_df, title_row, header_row, subset_df, empty_row])


    # Condition 1: MSSA
    add_section(MSSA, df[df[RESULTS] == MSSA])

    # Condition 2: MRSA PVL POSITIVE SSTI
    add_section(MRSA_POSITIVE_SSTI, df[(df[RESULTS] == MRSA) & (df[PVL] == POSITIVE) & (df[SPECIMEN_FINAL] == SSTI)])

    # Condition 3: MRSA PVL POSITIVE BLOOD
    add_section(MRSA_POSITIVE_BLOOD, df[(df[RESULTS] == MRSA) & (df[PVL] == POSITIVE) & (df[SPECIMEN_FINAL] == BLOOD)])

    # Condition 4: MRSA PVL NEGATIVE SSTI
    add_section(MRSA_NEGATIVE_SSTI, df[(df[RESULTS] == MRSA) & (df[PVL] == NEGATIVE) & (df[SPECIMEN_FINAL] == SSTI)])

    # Condition 5: MRSA PVL NEGATIVE BLOOD
    add_section(MRSA_NEGATIVE_BLOOD, df[(df[RESULTS] == MRSA) & (df[PVL] == NEGATIVE) & (df[SPECIMEN_FINAL] == BLOOD)])

    return final_df


def determine_result(row):
    '''
    if RT-mecA Result or RT-mecA Result positive = MRSA, if both negative = MSSA
    '''
    if row["RT-mecA Result"] == POSITIVE or row["RT-mecC Result"] == POSITIVE:
        return MRSA

    return MSSA


def style_excel(output_path, titles, data):
    wb = load_workbook(output_path)
    ws = wb.active

    # Ajustement dynamique de la largeur des colonnes
    column_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                col_letter = get_column_letter(cell.column)
                column_widths[col_letter] = max(column_widths.get(col_letter, 0), len(str(cell.value)))

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = min(width + 5, 50)

 
    bold_font = Font(bold=True)
    color_fill = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid")

    # Add value after each title
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        first_cell = row[0]
        title = first_cell.value

        if title in titles:
            for cell in row:
                cell.font = bold_font

            # Colorer la ligne suivante (en-tête)
            header_row = ws[row_idx + 1]  # Ligne sous le titre
            for cell in header_row:
                cell.fill = color_fill

            # Ajout des 3 valeurs avec des couleurs pastel
            title_column = first_cell.column

            if title in data:
                ws.cell(row=row_idx, column=title_column + 1, value=data[first_cell.value]['needed']).fill = GREEN_FILL
                ws.cell(row=row_idx, column=title_column + 2, value=data[first_cell.value]['found']).fill = YELLOW_FILL
                ws.cell(row=row_idx, column=title_column + 3, value=data[first_cell.value]['added']).fill = ORANGE_FILL

    wb.save(output_path)

def hightlight_added_rows(file_path, rows_added):
    """
    Processes an Excel file to find tables under specific titles and highlight extra rows.
    :param file_path: Path to the Excel file.
    :param rows_added: Dictionary {title_table: rows_needed} defining row limits per table.
    """
    wb = load_workbook(file_path)
    ws = wb.active

    df = pd.DataFrame(ws.values)

    for index, row in df.iterrows():
        title = row[0]

        if title in TITLES and title in rows_added:
            row_limit = rows_added[title]

            table_start = index + 1  # Table starts below title

            # Find the table end (empty row marks the end)
            table_end = table_start

            while table_end < len(df) and any(pd.notna(df.iloc[table_end])):
                table_end += 1

            table_size = table_end - table_start  # Number of rows in table

            # Highlight rows that exceed the row_limit
            if table_size > row_limit:
                for i in range(table_start + row_limit, table_end):
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=i + 1, column=col).fill = ORANGE_FILL  # OpenPyXL uses 1-based indexing

    wb.save(file_path)

 
def select_specimens(df, output_file, num_of_specimen=30):
    mrsa_negative_ssti_needed_before_completion = 0
    mrsa_negative_blood_needed_before_completion = 0
    mssa_needed_before_completion = 0

    # Create the SPECIMEN_FINAL column
    specimen_values_to_replace = ['Wound', 'Abscess', 'Surgery wound', 'Skin', 'Skin ulcer', 'Ear',
                                  'Ear L', 'Ear R', 'Elbow', 'Elbow L', 'Elbow R', 'Nose',
                                  'Lesion', 'Eye', 'Nasal wash', 'Navel', 'Perineum']

    # Create the SPECIMEN_FINAL column with the same values as 'Specimen'
    df[SPECIMEN_FINAL] = df[SPECIMEN]

    # Replace the specified values with SSTI in SPECIMEN_FINAL
    df[SPECIMEN_FINAL] = df[SPECIMEN_FINAL].replace(specimen_values_to_replace, SSTI)

    # Fix if RT-mecA Result or RT-mecA Result positive = MRSA, if both negative = MSSA
    df[RESULTS] = df.apply(determine_result, axis=1)

    # Creating the 'results' directory if it doesn't exist
    os.makedirs("results", exist_ok=True)

    # Ensure ORIGINATOR column is of type string
    df[ORIGINATOR] = df[ORIGINATOR].astype(str)

    # We need to take all the PVL Positive
    # The remaining will be splitten 50%, 25%, 25%
    match_criteria = (
        (df[RESULTS] == MRSA) &
        (df[PVL] == POSITIVE) &
        (df[SPECIMEN_FINAL].isin([SSTI, BLOOD]))

    )

    matched_df = df[match_criteria]
    matched_count = len(matched_df)
    remaining_needed = num_of_specimen - matched_count

    display_and_log(f"[+] Found {matched_count} matching MRSA PVL positive specimens. Need {remaining_needed} more.")

    remaining_df = df[(df[SPECIMEN_FINAL].isin([SSTI, BLOOD])) & ~match_criteria]

    ssti_criteria = (remaining_df[RESULTS] == MRSA) & (remaining_df[PVL] == NEGATIVE) & (remaining_df[SPECIMEN_FINAL] == SSTI)
    blood_criteria = (remaining_df[RESULTS] == MRSA) & (remaining_df[PVL] == NEGATIVE) & (remaining_df[SPECIMEN_FINAL] == BLOOD)

    # For MSSA, first select PVL positive, then fill with PVL negative if needed
    mssa_positive_criteria = (remaining_df[RESULTS] == MSSA) & (remaining_df[PVL] == POSITIVE) & (remaining_df[SPECIMEN_FINAL].isin([SSTI, BLOOD]))
    mssa_negative_criteria = (remaining_df[RESULTS] == MSSA) & (remaining_df[PVL] == NEGATIVE) & (remaining_df[SPECIMEN_FINAL].isin([SSTI, BLOOD]))

    mrsa_negative_ssti_needed  = math.ceil((remaining_needed * 0.5))
    mrsa_negative_blood_needed = math.ceil((remaining_needed * 0.25))
    mssa_needed             = math.ceil((remaining_needed * 0.25))

    display_and_log(f"[+] Total Needed : ")
    display_and_log(f"    - MRSA PVL NEGATIVE SSTI  :  {mrsa_negative_ssti_needed}.")
    display_and_log(f"    - MRSA PVL NEGATIVE BLOOD :  {mrsa_negative_blood_needed}.")
    display_and_log(f"    - MSSA PVL NEG / POS      : {mssa_needed}.")

    def select_by_originator(df, num_needed):
        if num_needed <= 0:
            return pd.DataFrame()

        originator_counts = df[ORIGINATOR].value_counts(normalize=True)
        selected_rows = []

        for originator, fraction in originator_counts.items():
            num_from_originator = max(1, int(round(fraction * num_needed)))
            subset = df[df[ORIGINATOR] == originator].sample(n=min(num_from_originator, len(df[df[ORIGINATOR] == originator])), random_state=42)
            selected_rows.append(subset)

        return pd.concat(selected_rows)[:num_needed]

    selected_ssti  = select_by_originator(remaining_df[ssti_criteria], mrsa_negative_ssti_needed)
    selected_blood = select_by_originator(remaining_df[blood_criteria], mrsa_negative_blood_needed)

    '''
    If found less data than needed from table (ex: if fetched 5 on 10 needed)
        -> if found less MRSA-NEG-SSTI : complete by more MRSA-NEG-BLOOD
        -> if found less MRSA-NEG-BLOOD : complete by more MRSA-NEG-SSTI
        -> the remaining to reach num_of_specimen (default 30) will be completed by MSSA
    '''
    ssti_fetched_len_on_start  = len(selected_ssti)
    blood_fetched_len_on_start = len(selected_blood)

    is_found_less_ssti_than_needed = mrsa_negative_ssti_needed > ssti_fetched_len_on_start
    is_found_less_blood_than_needed = mrsa_negative_blood_needed > blood_fetched_len_on_start

    if is_found_less_ssti_than_needed and not is_found_less_blood_than_needed:
        display_and_log(f"[!] MRSA-NEG-SSTI : Found {ssti_fetched_len_on_start} on {mrsa_negative_ssti_needed} needed : Completion with more MRSA-NEG-BLOOD ")

        mrsa_negative_blood_needed_before_completion = mrsa_negative_blood_needed
        mrsa_negative_blood_complement = mrsa_negative_ssti_needed - ssti_fetched_len_on_start
        mrsa_negative_blood_needed += mrsa_negative_blood_complement

        selected_blood  = select_by_originator(remaining_df[blood_criteria], mrsa_negative_blood_needed)

    elif not is_found_less_ssti_than_needed and is_found_less_blood_than_needed:
        display_and_log(f"[!] MRSA-NEG-SSTI : Found {ssti_fetched_len_on_start} on {mrsa_negative_ssti_needed} needed : Completion with more MRSA-NEG-BLOOD ")

        mrsa_negative_ssti_needed_before_completion = mrsa_negative_ssti_needed
        mrsa_negative_ssti_complement = mrsa_negative_blood_needed - blood_fetched_len_on_start
        mrsa_negative_ssti_needed += mrsa_negative_ssti_complement

        selected_ssti  = select_by_originator(remaining_df[ssti_criteria], mrsa_negative_ssti_needed)

    df_selection = pd.concat([matched_df, selected_ssti, selected_blood])

    '''
    If we have remaining, i.e, we need more specimen to reach num_of_specimen (default 30), we fetch from MSSA
        - First the PVL+ (if found some)
        - Second the PVL-
    '''

    if len(df_selection) < num_of_specimen:
        mssa_needed_before_completion = mssa_needed
        mssa_needed = num_of_specimen - len(df_selection)
        selected_mssa_positive = select_by_originator(remaining_df[mssa_positive_criteria], min(mssa_needed, len(remaining_df[mssa_positive_criteria])))

        # Update remaining MSSA needed after selecting PVL positive
        remaining_mssa_needed = mssa_needed - len(selected_mssa_positive)

        # If more MSSA specimens are needed, select PVL negative ones
        selected_mssa_negative = select_by_originator(remaining_df[mssa_negative_criteria], remaining_mssa_needed)

        display_and_log(f"[+] MSSA needed (neede + completion) total {mssa_needed}: ")
        display_and_log(f"    - PVL + : Found {len(selected_mssa_positive)}.")
        display_and_log(f"    - PVL - : Found {len(selected_mssa_negative)}.")

        df_selection = pd.concat([matched_df, selected_ssti, selected_blood, selected_mssa_positive, selected_mssa_negative])

    df_ordered = order_the_table(df_selection)
    df_ordered.to_excel(output_file, index=False, header=False)

    '''
    Add data next to each title table to indicate what needed found, completed
    '''
    mssa_len = len(selected_mssa_positive) + len(selected_mssa_negative)

    data_mssa_needed = mssa_needed_before_completion
    data_mssa_found = mssa_needed_before_completion if mssa_needed_before_completion < mssa_len else mssa_len
    data_mssa_added = mssa_len - mssa_needed_before_completion if mssa_needed_before_completion > 0 else 0

 
    ssti_len = len(selected_ssti)

    data_mrsa_negative_ssti_needed =  mrsa_negative_ssti_needed if mrsa_negative_ssti_needed_before_completion == 0 else mrsa_negative_ssti_needed_before_completion
    data_mrsa_negative_ssti_found  = ssti_len if mrsa_negative_ssti_needed_before_completion == 0 else ssti_fetched_len_on_start
    data_mrsa_negative_ssti_added = ssti_len - mrsa_negative_ssti_needed_before_completion if mrsa_negative_ssti_needed_before_completion > 0 else 0

    blood_len = len(selected_blood)
    data_mrsa_negative_blood_needed = mrsa_negative_blood_needed if mrsa_negative_blood_needed_before_completion == 0 else mrsa_negative_blood_needed_before_completion
    data_mrsa_negative_blood_found =  blood_len if mrsa_negative_blood_needed_before_completion == 0 else blood_fetched_len_on_start
    data_mrsa_negative_blood_added = blood_len - mrsa_negative_blood_needed_before_completion if mrsa_negative_blood_needed_before_completion > 0 else 0

    data = {
        MSSA: {
            'needed': f'Needed {data_mssa_needed}',
            'found': f'Found {data_mssa_found}',
            'added': f'Added {data_mssa_added} to reach {num_of_specimen}',
            },

        MRSA_NEGATIVE_SSTI: {
            'needed': f'Needed {data_mrsa_negative_ssti_needed}' ,
            'found': f'Found {data_mrsa_negative_ssti_found}',
            'added': f'Added {data_mrsa_negative_ssti_added} to reach {num_of_specimen}',

        },
        MRSA_NEGATIVE_BLOOD: {
            'needed': f'Needed {data_mrsa_negative_blood_needed}',
            'found': f'Found {data_mrsa_negative_blood_found}',
            'added': f'Added {data_mrsa_negative_blood_added} to reach {num_of_specimen}',

            }

    }

    style_excel(output_file, TITLES, data)

    rows_added = {}

    if mssa_needed_before_completion > 0:
        rows_added[MSSA] = mssa_needed_before_completion + 1 

    if mrsa_negative_blood_needed_before_completion > 0:
        rows_added[MRSA_NEGATIVE_BLOOD] = mrsa_negative_blood_needed_before_completion + 1

    if mrsa_negative_ssti_needed_before_completion > 0:
        rows_added[MRSA_NEGATIVE_SSTI] = mrsa_negative_ssti_needed_before_completion + 1


    if rows_added:   
        display_and_log(f"[+] Highlighting complementions rows.")
        hightlight_added_rows(output_file, rows_added)

 
    display_and_log(f"[+] File saved successfully to {output_file}.")


def main():
    parser = argparse.ArgumentParser(description="Select specimens to freeze from an input file.")
    parser.add_argument("--input_file", type=str, help="Path to input CSV or Excel file")
    parser.add_argument("--sheetname", required=False, type=str, help="sheetname")
    parser.add_argument("--num_of_specimen", type=int, default=30, help="num fof specimen")

    args = parser.parse_args()

    num_of_specimen = args.num_of_specimen

    setup_logs()

    display_and_log(f"\n\n[>>>] Generate the freeze bact file ({num_of_specimen} specimens)\n")

    df = load_data(args.input_file, args.sheetname)

    timestamp = datetime.now().strftime("%y%m%d_%H%M%S")
    output_filename = f"bact_to_freeze__{timestamp}.xlsx"
    
    select_specimens(df, output_filename, num_of_specimen)

    display_and_log(f"\n[---] The End ")


if __name__ == "__main__":

    main()
